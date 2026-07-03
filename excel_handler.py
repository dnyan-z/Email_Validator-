"""
excel_handler.py
----------------
Reads email data from the input Excel file and writes validation
results back to the output files.

Expected input columns (case-insensitive):
  • Mail ID     – the email address
  • Mail Status – optional existing status (will be overwritten)

Output adds:
  • Mail Status      – validation status code
  • Validation Reason – human-readable explanation
  • SMTP Response    – raw SMTP code + message
"""

import os
from collections import Counter

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill

from config import (
    FAILED_EXCEL,
    OUTPUT_DIR,
    OUTPUT_EXCEL,
    STATUS_VALID,
)
from logger import get_logger

log = get_logger(__name__)

# ── Colour palette for the status column ─────────────────────────────────────
_STATUS_COLOURS = {
    "VALID":             "C6EFCE",  # green
    "INVALID_FORMAT":    "FFEB9C",  # yellow
    "INVALID_DOMAIN":    "FFC7CE",  # red-light
    "INVALID_MAILBOX":   "FFC7CE",  # red-light
    "ACCESS_DENIED":     "FFCC99",  # orange
    "CATCH_ALL":         "BDD7EE",  # blue-light
    "TEMPORARY_FAILURE": "E2EFDA",  # grey-green
    "UNKNOWN":           "D9D9D9",  # grey
}


def read_emails(filepath: str) -> list[str]:
    """
    Read email addresses from an Excel file.

    Looks for a column whose name contains 'mail id' (case-insensitive).
    Falls back to the first column if no match is found.

    Parameters
    ----------
    filepath : str
        Path to the .xlsx input file.

    Returns
    -------
    list[str]
        All non-empty email strings found in the column.
    """
    log.info("Reading input file: %s", filepath)
    df = pd.read_excel(filepath, dtype=str)

    # Find the email column, but fall back safely if headers are numeric or mixed.
    col = _find_column(df, "mail id")
    if col is None:
        col = df.columns[0]

    emails = _extract_column_values(df, col)
    log.info("Using column '%s' for email addresses", _column_text(col))
    log.info("Found %d email addresses", len(emails))
    return emails


def write_results(results: list[dict]) -> None:
    """
    Write validation results to the output Excel file with colour-coding.

    Parameters
    ----------
    results : list[dict]
        Each dict must have keys: email, status, reason, smtp_response.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    df = _build_output_dataframe(results)

    wb = _build_workbook(df, sheet_name="Validation Results")
    _add_summary_sheet(wb, results)

    _save_workbook_resiliently(
        wb,
        target_path=OUTPUT_EXCEL,
        purpose="Results",
    )


def write_failed_emails(results: list[dict]) -> None:
    """
    Write only the failed / invalid email results to a separate file.

    Parameters
    ----------
    results : list[dict]
        Full results list; this function filters out VALID emails.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    failed = [r for r in results if r.get("status") != STATUS_VALID]
    if not failed:
        log.info("No failed emails – skipping failed_emails.xlsx")
        return

    df = _build_output_dataframe(failed)

    wb = _build_workbook(df, sheet_name="Failed Emails")
    _add_summary_sheet(wb, failed)

    _save_workbook_resiliently(
        wb,
        target_path=FAILED_EXCEL,
        purpose="Failed emails",
    )


# ── Private helpers ───────────────────────────────────────────────────────────

def _save_workbook_resiliently(wb: Workbook, target_path: str, purpose: str) -> None:
    """
    Save workbook to `target_path`, working around Windows file locks (Excel open)
    by saving to a temp file first and then atomically replacing.

    If replace fails due to PermissionError, writes to a timestamped fallback file.
    """
    import time
    import uuid

    target_dir = os.path.dirname(target_path) or "."
    os.makedirs(target_dir, exist_ok=True)

    base_name = os.path.basename(target_path)
    tmp_path = os.path.join(
        target_dir,
        f".{base_name}.tmp_{os.getpid()}_{uuid.uuid4().hex[:8]}",
    )

    # 1) Save to temp first.
    wb.save(tmp_path)

    # 2) Try atomic replace.
    try:
        os.replace(tmp_path, target_path)
        log.info("%s saved → %s", purpose, target_path)
        return
    except PermissionError:
        # 3) Fallback: write to timestamped filename instead of crashing.
        ts = time.strftime("%Y%m%d_%H%M%S")
        fallback_path = os.path.join(
            target_dir, f"{os.path.splitext(base_name)[0]}_{ts}.xlsx"
        )
        try:
            os.replace(tmp_path, fallback_path)
        except PermissionError:
            log.exception(
                "%s could not overwrite locked file and fallback also failed (PermissionError). target=%s fallback=%s",
                purpose,
                target_path,
                fallback_path,
            )
            raise
        log.warning(
            "%s could not overwrite locked file (PermissionError). Wrote fallback → %s",
            purpose,
            fallback_path,
        )
        return

def _find_column(df: pd.DataFrame, keyword: str) -> str | None:
    """Return the first column name whose lowercase form contains keyword."""
    for col in df.columns:
        col_name = _column_text(col)
        if keyword in col_name.lower():
            return col
    return None


def _column_text(column: object) -> str:
    """Convert any Excel column label to a safe string for comparison."""
    if column is None:
        return ""
    return str(column).strip()


def _extract_column_values(df: pd.DataFrame, column: object) -> list[str]:
    """Return cleaned string values from a column label or positional fallback."""
    if column in df.columns:
        series = df[column]
    else:
        series = df.iloc[:, 0]

    values = series.dropna().astype(str).map(str.strip).tolist()
    return [value for value in values if value]


def _build_workbook(df: pd.DataFrame, sheet_name: str) -> Workbook:
    """
    Convert a DataFrame into a formatted openpyxl Workbook.

    Parameters
    ----------
    df : pd.DataFrame
        Data to write (already has correct column names).
    sheet_name : str
        Name for the worksheet.

    Returns
    -------
    Workbook
        Ready-to-save workbook with header formatting and row colours.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ── Headers ───────────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill  = PatternFill("solid", start_color="2E75B6")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align

    # ── Data rows ─────────────────────────────────────────────────────────────
    status_col_idx = list(df.columns).index("Mail Status") + 1
    risk_col_idx = list(df.columns).index("Risk Score") + 1
    deliver_col_idx = list(df.columns).index("Deliverability Score") + 1

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")

        # Colour the status cell
        status_value = ws.cell(row=row_idx, column=status_col_idx).value or ""
        colour = _STATUS_COLOURS.get(status_value, "FFFFFF")
        ws.cell(row=row_idx, column=status_col_idx).fill = PatternFill(
            "solid", start_color=colour
        )

    # ── Workbook usability features ──────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto width with a cap to keep files readable.
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col_letter].width = min(60, max(12, max_len + 2))

    ws.row_dimensions[1].height = 20

    # Conditional formatting for risk and deliverability.
    ws.conditional_formatting.add(
        f"{ws.cell(row=2, column=risk_col_idx).coordinate}:{ws.cell(row=ws.max_row, column=risk_col_idx).coordinate}",
        CellIsRule(operator="greaterThanOrEqual", formula=["70"], stopIfTrue=True, fill=PatternFill("solid", start_color="FFC7CE")),
    )
    ws.conditional_formatting.add(
        f"{ws.cell(row=2, column=deliver_col_idx).coordinate}:{ws.cell(row=ws.max_row, column=deliver_col_idx).coordinate}",
        CellIsRule(operator="lessThan", formula=["50"], stopIfTrue=True, fill=PatternFill("solid", start_color="FFEB9C")),
    )

    return wb


def _build_output_dataframe(results: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(results)
    for column in [
        "normalized_email",
        "domain",
        "provider_classification",
        "is_role_account",
        "is_disposable",
        "is_free_provider",
        "smtp_code",
        "smtp_message",
        "catch_all_result",
        "risk_score",
        "confidence_score",
        "deliverability_score",
        "validation_duration_ms",
        "reason_code",
        "email",
        "status",
        "reason",
        "smtp_response",
    ]:
        if column not in df.columns:
            df[column] = ""

    df = df[
        [
            "email",
            "normalized_email",
            "domain",
            "provider_classification",
            "status",
            "reason",
            "reason_code",
            "is_role_account",
            "is_disposable",
            "is_free_provider",
            "smtp_code",
            "smtp_message",
            "smtp_response",
            "catch_all_result",
            "risk_score",
            "confidence_score",
            "deliverability_score",
            "validation_duration_ms",
        ]
    ]
    df.columns = [
        "Mail ID",
        "Normalized Email",
        "Domain",
        "Provider",
        "Mail Status",
        "Validation Reason",
        "Reason Codes",
        "Role Account",
        "Disposable",
        "Free Provider",
        "SMTP Code",
        "SMTP Message",
        "SMTP Response",
        "Catch-All Result",
        "Risk Score",
        "Confidence",
        "Deliverability Score",
        "Validation Time (ms)",
    ]
    return df


def _add_summary_sheet(wb: Workbook, results: list[dict]) -> None:
    ws = wb.create_sheet(title="Summary")
    statuses = Counter(r.get("status", "UNKNOWN") for r in results)
    domains = Counter(r.get("domain", "") for r in results if r.get("domain"))

    ws.append(["Metric", "Value"])
    ws.append(["Total Emails", len(results)])
    ws.append(["Valid", statuses.get("VALID", 0)])
    ws.append(["Invalid", statuses.get("INVALID_FORMAT", 0) + statuses.get("INVALID_DOMAIN", 0) + statuses.get("INVALID_MAILBOX", 0)])
    ws.append(["Catch-all", statuses.get("CATCH_ALL", 0)])
    ws.append(["Temporary failures", statuses.get("TEMPORARY_FAILURE", 0)])
    ws.append(["Unknown", statuses.get("UNKNOWN", 0)])
    ws.append([])
    ws.append(["Top Domains", "Count"])
    for domain, count in domains.most_common(10):
        ws.append([domain, count])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col_letter].width = min(40, max(12, max_len + 2))
